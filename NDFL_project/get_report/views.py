from django.http import HttpResponse
from django.views.generic.edit import FormView
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from .forms import ReportForm

class ExcelReportView(FormView):
    form_class = ReportForm
    template_name = 'report.html'
    
    def form_valid(self, form):
        # исходный файл
        excel_file = form.cleaned_data['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
        # создание файла для отчета
        new_wb = Workbook()
        new_ws = new_wb.active
        # копирование столбцов их исходных данных
        column_a = [cell.value for cell in ws["A"]]
        for i, value in enumerate(column_a, start=1):
            new_ws.cell(row=i, column=1, value=value)
        new_ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        column_b = [cell.value for cell in ws["B"]]
        for i, value in enumerate(column_b, start=1):
            new_ws.cell(row=i, column=2, value=value)
        new_ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        column_c = [cell.value for cell in ws["E"]]
        for i, value in enumerate(column_c, start=1):
            if value:
                new_ws.cell(row=i, column=3, value=value)
            else:
                new_ws.cell(row=i, column=3, value='')
        new_ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        column_d = [cell.value for cell in ws["F"]]
        for i, value in enumerate(column_d, start=1):
            if value:
                new_ws.cell(row=i, column=4, value=value)
            else:
                new_ws.cell(row=i, column=4, value='')
        # формирование столбца 'Исчислено всего по формуле'
        new_ws['E2'] = 'Исчислено всего по формуле'
        new_ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
        # формирование столбца 'Отклонения'
        new_ws['F1'] = 'Отклонения'
        new_ws.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)
        # расчет столбцов 'Исчислено всего по формуле' и 'Отклонения'
        for i, value in enumerate(column_c, start=1):
            if i > 3:
                if value:
                    if value < 5000000:
                        num = float(value)/100*13
                        diff = new_ws[f'D{i}'].value - num
                        new_ws[f'E{i}'] = int(num + (0.5 if num > 0 else -0.5))
                        new_ws[f'F{i}'] = int(diff + (0.5 if num > 0 else -0.5))
                    else:
                        num = float(value)/100*15
                        diff = new_ws[f'D{i}'].value - num
                        new_ws[f'E{i}'] = int(num + (0.5 if num > 0 else -0.5))
                        new_ws[f'F{i}'] = int(diff + (0.5 if num > 0 else -0.5))
                else:
                    new_ws[f'E{i}'] = ''
                    new_ws[f'F{i}'] = 0
        # сортировка списка по убыванию в столбце 'Отклонения'
        data = list(new_ws.values)
        sorted_data = sorted(data[3:(new_ws.max_row-1)], key=lambda x: x[5], reverse=True)
        # запись отсортированных данных обратно в лист
        for row_index, row_data in enumerate(sorted_data):
            for col_index, value in enumerate(row_data):
                new_ws.cell(row=row_index+4, column=col_index+1, value=value)
        # изменение фона ячеек
        column_f = [cell.value for cell in new_ws["F"]]
        for i, value in enumerate(column_f, start=1):
            if i > 3 and i < new_ws.max_row:
                if value == 0:
                    new_ws[f'F{i}'].fill = PatternFill('solid', fgColor="00ff00")
                else:
                    new_ws[f'F{i}'].fill = PatternFill('solid', fgColor="ff0000")
        # сохранение файла
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=report.xlsx'
        new_wb.save(response)
        return response